"""
Rutas para exportación de datos
"""
import io
import json
from datetime import datetime
from flask import Blueprint, request, Response
from flask_jwt_extended import jwt_required
from sqlalchemy import text

from models import db, Admin, Alumno, Carrera, Materia, Calificacion, NotaRemision, PracticaProfesional
from utils.decorators import admin_required

export_bp = Blueprint('export', __name__)


@export_bp.route('/sql', methods=['GET'])
@admin_required
def export_sql():
    """
    Genera un dump SQL completo de todas las tablas
    """
    output = io.StringIO()
    
    # Header con información
    output.write("-- ============================================\n")
    output.write(f"-- Portal FV - Dump SQL Completo\n")
    output.write(f"-- Generado: {datetime.utcnow().isoformat()}\n")
    output.write("-- ============================================\n\n")
    
    # Desactivar foreign keys para poder hacer truncate
    output.write("PRAGMA foreign_keys = OFF;\n\n")
    
    # Exportar Admins (sin passwords para seguridad)
    output.write("-- ============================================\n")
    output.write("-- TABLA: admins\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS admins (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    username TEXT UNIQUE NOT NULL,\n")
    output.write("    email TEXT UNIQUE NOT NULL,\n")
    output.write("    password_hash TEXT NOT NULL,\n")
    output.write("    nombre TEXT NOT NULL,\n")
    output.write("    created_at TIMESTAMP,\n")
    output.write("    updated_at TIMESTAMP\n")
    output.write(");\n\n")
    
    admins = Admin.query.all()
    for admin in admins:
        output.write(f"INSERT INTO admins (id, username, email, password_hash, nombre, created_at, updated_at) VALUES (\n")
        output.write(f"    {admin.id},\n")
        output.write(f"    '{admin.username}',\n")
        output.write(f"    '{admin.email}',\n")
        output.write(f"    '{admin.password_hash}',\n")
        output.write(f"    '{admin.nombre}',\n")
        output.write(f"    '{admin.created_at.isoformat() if admin.created_at else ''}',\n")
        output.write(f"    '{admin.updated_at.isoformat() if admin.updated_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Carreras
    output.write("-- ============================================\n")
    output.write("-- TABLA: carreras\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS carreras (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    nombre TEXT NOT NULL,\n")
    output.write("    codigo TEXT UNIQUE NOT NULL,\n")
    output.write("    descripcion TEXT,\n")
    output.write("    activa INTEGER DEFAULT 1,\n")
    output.write("    created_at TIMESTAMP\n")
    output.write(");\n\n")
    
    carreras = Carrera.query.all()
    for carrera in carreras:
        descripcion = carrera.descripcion.replace("'", "''") if carrera.descripcion else ''
        output.write(f"INSERT INTO carreras (id, nombre, codigo, descripcion, activa, created_at) VALUES (\n")
        output.write(f"    {carrera.id},\n")
        output.write(f"    '{carrera.nombre}',\n")
        output.write(f"    '{carrera.codigo}',\n")
        output.write(f"    '{descripcion}',\n")
        output.write(f"    {1 if carrera.activa else 0},\n")
        output.write(f"    '{carrera.created_at.isoformat() if carrera.created_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Materias
    output.write("-- ============================================\n")
    output.write("-- TABLA: materias\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS materias (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    carrera_id INTEGER REFERENCES carreras(id),\n")
    output.write("    nombre TEXT NOT NULL,\n")
    output.write("    codigo TEXT NOT NULL,\n")
    output.write("    creditos INTEGER DEFAULT 0,\n")
    output.write("    created_at TIMESTAMP\n")
    output.write(");\n\n")
    
    materias = Materia.query.all()
    for materia in materias:
        output.write(f"INSERT INTO materias (id, carrera_id, nombre, codigo, creditos, created_at) VALUES (\n")
        output.write(f"    {materia.id},\n")
        output.write(f"    {materia.carrera_id},\n")
        output.write(f"    '{materia.nombre}',\n")
        output.write(f"    '{materia.codigo}',\n")
        output.write(f"    {materia.creditos},\n")
        output.write(f"    '{materia.created_at.isoformat() if materia.created_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Alumnos
    output.write("-- ============================================\n")
    output.write("-- TABLA: alumnos\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS alumnos (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    numero_control TEXT UNIQUE NOT NULL,\n")
    output.write("    nombre TEXT NOT NULL,\n")
    output.write("    apellido_paterno TEXT NOT NULL,\n")
    output.write("    apellido_materno TEXT,\n")
    output.write("    email TEXT UNIQUE NOT NULL,\n")
    output.write("    password_hash TEXT NOT NULL,\n")
    output.write("    carrera_id INTEGER REFERENCES carreras(id),\n")
    output.write("    activo INTEGER DEFAULT 1,\n")
    output.write("    fecha_registro DATE,\n")
    output.write("    created_at TIMESTAMP\n")
    output.write(");\n\n")
    
    alumnos = Alumno.query.all()
    for alumno in alumnos:
        fecha_registro = alumno.fecha_registro.isoformat() if alumno.fecha_registro else ''
        output.write(f"INSERT INTO alumnos (id, numero_control, nombre, apellido_paterno, apellido_materno, email, password_hash, carrera_id, activo, fecha_registro, created_at) VALUES (\n")
        output.write(f"    {alumno.id},\n")
        output.write(f"    '{alumno.numero_control}',\n")
        output.write(f"    '{alumno.nombre}',\n")
        output.write(f"    '{alumno.apellido_paterno}',\n")
        output.write(f"    '{alumno.apellido_materno or ''}',\n")
        output.write(f"    '{alumno.email}',\n")
        output.write(f"    '{alumno.password_hash}',\n")
        output.write(f"    {alumno.carrera_id},\n")
        output.write(f"    {1 if alumno.activo else 0},\n")
        output.write(f"    '{fecha_registro}',\n")
        output.write(f"    '{alumno.created_at.isoformat() if alumno.created_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Calificaciones
    output.write("-- ============================================\n")
    output.write("-- TABLA: calificaciones\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS calificaciones (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    alumno_id INTEGER REFERENCES alumnos(id),\n")
    output.write("    materia_id INTEGER REFERENCES materias(id),\n")
    output.write("    asistencia_1 INTEGER DEFAULT 0,\n")
    output.write("    asistencia_2 INTEGER DEFAULT 0,\n")
    output.write("    asistencia_3 INTEGER DEFAULT 0,\n")
    output.write("    asistencia_4 INTEGER DEFAULT 0,\n")
    output.write("    asistencia_5 INTEGER DEFAULT 0,\n")
    output.write("    practica_1 REAL DEFAULT 0,\n")
    output.write("    practica_2 REAL DEFAULT 0,\n")
    output.write("    calificacion_final REAL DEFAULT 0,\n")
    output.write("    periodo TEXT,\n")
    output.write("    anio INTEGER,\n")
    output.write("    created_at TIMESTAMP,\n")
    output.write("    updated_at TIMESTAMP\n")
    output.write(");\n\n")
    
    calificaciones = Calificacion.query.all()
    for cal in calificaciones:
        output.write(f"INSERT INTO calificaciones (id, alumno_id, materia_id, asistencia_1, asistencia_2, asistencia_3, asistencia_4, asistencia_5, practica_1, practica_2, calificacion_final, periodo, anio, created_at, updated_at) VALUES (\n")
        output.write(f"    {cal.id},\n")
        output.write(f"    {cal.alumno_id},\n")
        output.write(f"    {cal.materia_id},\n")
        output.write(f"    {cal.asistencia_1},\n")
        output.write(f"    {cal.asistencia_2},\n")
        output.write(f"    {cal.asistencia_3},\n")
        output.write(f"    {cal.asistencia_4},\n")
        output.write(f"    {cal.asistencia_5},\n")
        output.write(f"    {cal.practica_1},\n")
        output.write(f"    {cal.practica_2},\n")
        output.write(f"    {cal.calificacion_final},\n")
        output.write(f"    '{cal.periodo}',\n")
        output.write(f"    {cal.anio},\n")
        output.write(f"    '{cal.created_at.isoformat() if cal.created_at else ''}',\n")
        output.write(f"    '{cal.updated_at.isoformat() if cal.updated_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Prácticas Profesionales
    output.write("-- ============================================\n")
    output.write("-- TABLA: practicas_profesionales\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS practicas_profesionales (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    alumno_id INTEGER REFERENCES alumnos(id),\n")
    output.write("    numero_practica INTEGER NOT NULL,\n")
    output.write("    nombre_empresa TEXT,\n")
    output.write("    fecha_inicio DATE,\n")
    output.write("    fecha_fin DATE,\n")
    output.write("    reporte_entregado INTEGER DEFAULT 0,\n")
    output.write("    validada INTEGER DEFAULT 0,\n")
    output.write("    observaciones TEXT,\n")
    output.write("    created_at TIMESTAMP,\n")
    output.write("    updated_at TIMESTAMP\n")
    output.write(");\n\n")
    
    practicas = PracticaProfesional.query.all()
    for p in practicas:
        fecha_inicio = p.fecha_inicio.isoformat() if p.fecha_inicio else ''
        fecha_fin = p.fecha_fin.isoformat() if p.fecha_fin else ''
        observaciones = p.observaciones.replace("'", "''") if p.observaciones else ''
        empresa = p.nombre_empresa.replace("'", "''") if p.nombre_empresa else ''
        
        output.write(f"INSERT INTO practicas_profesionales (id, alumno_id, numero_practica, nombre_empresa, fecha_inicio, fecha_fin, reporte_entregado, validada, observaciones, created_at, updated_at) VALUES (\n")
        output.write(f"    {p.id},\n")
        output.write(f"    {p.alumno_id},\n")
        output.write(f"    {p.numero_practica},\n")
        output.write(f"    '{empresa}',\n")
        output.write(f"    '{fecha_inicio}',\n")
        output.write(f"    '{fecha_fin}',\n")
        output.write(f"    {1 if p.reporte_entregado else 0},\n")
        output.write(f"    {1 if p.validada else 0},\n")
        output.write(f"    '{observaciones}',\n")
        output.write(f"    '{p.created_at.isoformat() if p.created_at else ''}',\n")
        output.write(f"    '{p.updated_at.isoformat() if p.updated_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n")
    
    # Exportar Notas de Remisión
    output.write("-- ============================================\n")
    output.write("-- TABLA: notas_remision\n")
    output.write("-- ============================================\n")
    output.write("CREATE TABLE IF NOT EXISTS notas_remision (\n")
    output.write("    id INTEGER PRIMARY KEY AUTOINCREMENT,\n")
    output.write("    alumno_id INTEGER REFERENCES alumnos(id),\n")
    output.write("    concepto TEXT NOT NULL,\n")
    output.write("    monto REAL NOT NULL,\n")
    output.write("    fecha_emision DATE,\n")
    output.write("    pagada INTEGER DEFAULT 0,\n")
    output.write("    fecha_pago DATE,\n")
    output.write("    created_by_id INTEGER REFERENCES admins(id),\n")
    output.write("    created_at TIMESTAMP,\n")
    output.write("    updated_at TIMESTAMP\n")
    output.write(");\n\n")
    
    notas = NotaRemision.query.all()
    for n in notas:
        fecha_emision = n.fecha_emision.isoformat() if n.fecha_emision else ''
        fecha_pago = n.fecha_pago.isoformat() if n.fecha_pago else ''
        concepto = n.concepto.replace("'", "''")
        
        output.write(f"INSERT INTO notas_remision (id, alumno_id, concepto, monto, fecha_emision, pagada, fecha_pago, created_by_id, created_at, updated_at) VALUES (\n")
        output.write(f"    {n.id},\n")
        output.write(f"    {n.alumno_id},\n")
        output.write(f"    '{concepto}',\n")
        output.write(f"    {n.monto},\n")
        output.write(f"    '{fecha_emision}',\n")
        output.write(f"    {1 if n.pagada else 0},\n")
        output.write(f"    '{fecha_pago}',\n")
        output.write(f"    {n.created_by_id},\n")
        output.write(f"    '{n.created_at.isoformat() if n.created_at else ''}',\n")
        output.write(f"    '{n.updated_at.isoformat() if n.updated_at else ''}'\n")
        output.write(");\n")
    
    output.write("\n-- Reactivar foreign keys\n")
    output.write("PRAGMA foreign_keys = ON;\n")
    
    # Generar respuesta
    output.seek(0)
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    return Response(
        output.getvalue(),
        mimetype='application/sql',
        headers={
            'Content-Disposition': f'attachment; filename=portal_fv_backup_{timestamp}.sql'
        }
    )


@export_bp.route('/excel', methods=['GET'])
@admin_required
def export_excel():
    """
    Genera un archivo Excel con todos los datos
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='008a8a', end_color='008a8a', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # Hoja: Carreras
        ws_carreras = wb.active
        ws_carreras.title = 'Carreras'
        ws_carreras.append(['ID', 'Nombre', 'Código', 'Descripción', 'Activa'])
        style_header(ws_carreras)
        for c in Carrera.query.all():
            ws_carreras.append([c.id, c.nombre, c.codigo, c.descripcion or '', 'Sí' if c.activa else 'No'])
        
        # Hoja: Materias
        ws_materias = wb.create_sheet('Materias')
        ws_materias.append(['ID', 'Carrera', 'Nombre', 'Código', 'Créditos'])
        style_header(ws_materias)
        for m in Materia.query.all():
            ws_materias.append([m.id, m.carrera.nombre if m.carrera else '', m.nombre, m.codigo, m.creditos])
        
        # Hoja: Alumnos
        ws_alumnos = wb.create_sheet('Alumnos')
        ws_alumnos.append(['ID', 'No. Control', 'Nombre', 'Apellido Paterno', 'Apellido Materno', 'Email', 'Carrera', 'Activo'])
        style_header(ws_alumnos)
        for a in Alumno.query.all():
            ws_alumnos.append([
                a.id, a.numero_control, a.nombre, a.apellido_paterno,
                a.apellido_materno or '', a.email,
                a.carrera.nombre if a.carrera else '', 'Sí' if a.activo else 'No'
            ])
        
        # Hoja: Calificaciones
        ws_calificaciones = wb.create_sheet('Calificaciones')
        ws_calificaciones.append([
            'ID', 'Alumno', 'Materia', 'Asistencia 1', 'Asistencia 2', 'Asistencia 3',
            'Asistencia 4', 'Asistencia 5', 'Práctica 1', 'Práctica 2', 'Cal. Final', 'Período', 'Año'
        ])
        style_header(ws_calificaciones)
        for c in Calificacion.query.all():
            ws_calificaciones.append([
                c.id,
                c.alumno.nombre_completo if c.alumno else '',
                c.materia.nombre if c.materia else '',
                c.asistencia_1, c.asistencia_2, c.asistencia_3, c.asistencia_4, c.asistencia_5,
                c.practica_1, c.practica_2, c.calificacion_final,
                c.periodo or '', c.anio or ''
            ])
        
        # Hoja: Notas de Remisión
        ws_notas = wb.create_sheet('Notas Remisión')
        ws_notas.append(['ID', 'Alumno', 'Concepto', 'Monto', 'Fecha Emisión', 'Pagada', 'Fecha Pago'])
        style_header(ws_notas)
        for n in NotaRemision.query.all():
            ws_notas.append([
                n.id,
                n.alumno.nombre_completo if n.alumno else '',
                n.concepto, n.monto,
                n.fecha_emision.isoformat() if n.fecha_emision else '',
                'Sí' if n.pagada else 'No',
                n.fecha_pago.isoformat() if n.fecha_pago else ''
            ])
        
        # Hoja: Prácticas
        ws_practicas = wb.create_sheet('Prácticas')
        ws_practicas.append(['ID', 'Alumno', 'No. Práctica', 'Empresa', 'Fecha Inicio', 'Fecha Fin', 'Reporte', 'Validada'])
        style_header(ws_practicas)
        for p in PracticaProfesional.query.all():
            ws_practicas.append([
                p.id,
                p.alumno.nombre_completo if p.alumno else '',
                p.numero_practica, p.nombre_empresa or '',
                p.fecha_inicio.isoformat() if p.fecha_inicio else '',
                p.fecha_fin.isoformat() if p.fecha_fin else '',
                'Sí' if p.reporte_entregado else 'No',
                'Sí' if p.validada else 'No'
            ])
        
        # Ajustar anchos de columna
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=portal_fv_data_{timestamp}.xlsx'
            }
        )
        
    except ImportError:
        return jsonify({'error': 'openpyxl no está instalado'}), 500
    except Exception as e:
        return jsonify({'error': f'Error al generar Excel: {str(e)}'}), 500


@export_bp.route('/json', methods=['GET'])
@admin_required
def export_json():
    """
    Genera un archivo JSON con todos los datos
    """
    data = {
        'export_date': datetime.utcnow().isoformat(),
        'version': '1.0',
        'carreras': [c.to_dict() for c in Carrera.query.all()],
        'materias': [m.to_dict() for m in Materia.query.all()],
        'alumnos': [a.to_dict() for a in Alumno.query.all()],
        'calificaciones': [c.to_dict() for c in Calificacion.query.all()],
        'notas_remision': [n.to_dict() for n in NotaRemision.query.all()],
        'practicas': [p.to_dict() for p in PracticaProfesional.query.all()]
    }
    
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    return Response(
        json.dumps(data, indent=2, default=str),
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename=portal_fv_data_{timestamp}.json'
        }
    )
