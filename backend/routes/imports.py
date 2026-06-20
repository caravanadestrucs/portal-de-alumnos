"""
Rutas para Importación Masiva de Datos (CSV / XLSX)
Soporta: Alumnos, Calificaciones, Pagos, Carreras, Materias
"""
import csv
import io
import os
import re
import tempfile
from datetime import date, datetime
from werkzeug.utils import secure_filename

from flask import Blueprint, request, jsonify
from flask_jwt_extended import get_jwt

from models import db, Alumno, Carrera, Materia, Calificacion, NotaRemision
from utils.decorators import admin_required

imports_bp = Blueprint('imports', __name__)

# ============================================================
# CONSTANTES
# ============================================================

ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}
BATCH_SIZE = 500
MAX_CONTENT_LENGTH = 10 * 1024 * 1024  # 10 MB

# ============================================================
# HEADER ALIASES — mapeo de variantes a nombres canónicos
# ============================================================

HEADER_ALIASES = {
    'numero_control': [
        'numero_control', 'numero de control', 'no_control', 'no control',
        'no._control', 'num_control', 'num control', 'no', 'matricula',
        'numero_de_control', 'número de control', 'número_de_control',
        'numerocontrol',
    ],
    'nombre': [
        'nombre', 'nombres', 'nombre_s', 'name', 'nombre_del_alumno',
    ],
    'apellido_paterno': [
        'apellido_paterno', 'apellido paterno', 'apellido1', 'apellido_1',
        'primer_apellido', 'primer apellido', 'apellidopaterno',
    ],
    'apellido_materno': [
        'apellido_materno', 'apellido materno', 'apellido2', 'apellido_2',
        'segundo_apellido', 'segundo apellido', 'apellidomaterno',
    ],
    'email': [
        'email', 'e_mail', 'e-mail', 'mail', 'correo',
        'correo_electronico', 'correo electronico',
        'correo_electrónico', 'correo electrónico',
    ],
    'carrera': [
        'carrera', 'codigo_carrera', 'codigo carrera',
        'carrera_codigo', 'carrera codigo',
        'código carrera', 'código_carrera',
        'carrera_nombre', 'carrera nombre',
    ],
    'password': [
        'password', 'contraseña', 'contrasena', 'contrasenia',
        'pass', 'clave', 'passwd',
    ],
    'materia': [
        'materia', 'codigo_materia', 'codigo materia',
        'materia_codigo', 'materia codigo',
        'código materia', 'código_materia',
        'asignatura', 'materia_nombre', 'materia nombre',
    ],
    'calificacion_final': [
        'calificacion_final', 'calificacion final',
        'calificación final', 'calificación_final',
        'calificacion', 'calificación', 'calif',
        'nota_final', 'nota final', 'calif_final',
        'nota', 'promedio', 'calificacionfinal',
    ],
    'periodo': [
        'periodo', 'período', 'period', 'semestre',
        'periodo_', 'cuatrimestre',
    ],
    'anio': [
        'anio', 'año', 'year', 'annio', 'año_', 'anio_lectivo',
    ],
    'practica_1': [
        'practica_1', 'practica1', 'practica 1',
        'práctica 1', 'práctica_1', 'practica_01',
    ],
    'practica_2': [
        'practica_2', 'practica2', 'practica 2',
        'práctica 2', 'práctica_2', 'practica_02',
    ],
    'extra_1': [
        'extra_1', 'extra1', 'extra 1',
        'ordinario_1', 'ordinario1', 'recuperacion_1',
    ],
    'extra_2': [
        'extra_2', 'extra2', 'extra 2',
        'ordinario_2', 'ordinario2', 'recuperacion_2',
    ],
    'asistencia_1': [
        'asistencia_1', 'asistencia1', 'asistencia 1', 'asis_1',
    ],
    'asistencia_2': [
        'asistencia_2', 'asistencia2', 'asistencia 2', 'asis_2',
    ],
    'asistencia_3': [
        'asistencia_3', 'asistencia3', 'asistencia 3', 'asis_3',
    ],
    'asistencia_4': [
        'asistencia_4', 'asistencia4', 'asistencia 4', 'asis_4',
    ],
    'asistencia_5': [
        'asistencia_5', 'asistencia5', 'asistencia 5', 'asis_5',
    ],
    'concepto': [
        'concepto', 'descripcion', 'descripción', 'concept',
        'motivo', 'razon',
    ],
    'monto': [
        'monto', 'importe', 'cantidad', 'amount', 'total',
        'costo', 'adeudo', 'monto_total',
    ],
    'fecha_emision': [
        'fecha_emision', 'fecha_emisión', 'fecha emision',
        'fecha emisión', 'fecha_de_emision', 'fecha de emision',
        'emision', 'emisión', 'fecha_expedicion',
    ],
    'fecha_corte': [
        'fecha_corte', 'fecha corte', 'fecha_de_corte',
        'fecha de corte', 'corte', 'fecha_limite', 'fecha limite',
        'fecha_vencimiento',
    ],
    'pagada': [
        'pagada', 'pagado', 'estado_pago', 'estado pago',
        'estado', 'paid', 'pagada_', 'estado_de_pago',
    ],
    'fecha_pago': [
        'fecha_pago', 'fecha pago', 'fecha_de_pago',
        'fecha de pago', 'fecha_pagada', 'fecha pagada',
        'fecha_de_pagado',
    ],
    'codigo': [
        'codigo', 'código', 'code', 'clave', 'key',
        'codigo_carrera', 'código carrera', 'codigo_materia',
    ],
    'descripcion': [
        'descripcion', 'descripción', 'description', 'desc',
        'observaciones', 'notas', 'comentarios',
    ],
    'creditos': [
        'creditos', 'créditos', 'credits', 'horas',
        'num_creditos', 'no_creditos', 'numero_creditos',
    ],
    'activa': [
        'activa', 'activo', 'active', 'activa_', 'estado',
        'habilitada', 'habilitado', 'enabled',
    ],
}

# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

# Correo regex simple
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')


def _normalize_header(header):
    """
    Normaliza un header: limpia, lowercase, sin acentos,
    espacios → underscore.
    """
    if not header:
        return ''
    h = str(header).strip().lower()
    # Reemplazar acentos
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'ü': 'u', 'ñ': 'n',
    }
    for k, v in replacements.items():
        h = h.replace(k, v)
    # Remover caracteres especiales (excepto letras, números, espacios, _)
    h = re.sub(r'[^\w\s]', ' ', h)
    # Espacios múltiples → underscore
    h = re.sub(r'\s+', '_', h)
    # Underscores múltiples → uno solo
    h = re.sub(r'_+', '_', h)
    h = h.strip('_')
    return h


def _sanitize_value(val):
    """
    Sanitiza un valor para prevenir CSV injection.
    Si comienza con =, +, -, @ → prefija con comilla simple.
    """
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@'):
        return "'" + val
    return val


# Precomputar reverse lookup: variante normalizada → campo canónico
_ALIAS_TO_FIELD = {}
for canonical, variants in HEADER_ALIASES.items():
    for variant in variants:
        _ALIAS_TO_FIELD[_normalize_header(variant)] = canonical
    # También mapear el nombre canónico a sí mismo
    _ALIAS_TO_FIELD[canonical] = canonical


def _detect_file_type(filename):
    """Valida que la extensión sea .csv o .xlsx. Retorna la extensión."""
    if not filename or '.' not in filename:
        raise ValueError("El archivo no tiene extensión")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Formato no soportado: '{ext}'. Use .csv o .xlsx")
    return ext


def _read_file(file_storage):
    """
    Lee un archivo CSV o XLSX desde un FileStorage de Flask.

    Returns:
        dict con:
          - headers: list[str] — headers normalizados
          - rows: list[list[str]] — filas como listas de strings
          - total_rows: int
    """
    ext = _detect_file_type(file_storage.filename)
    file_storage.seek(0)

    headers = []
    rows = []

    if ext == '.csv':
        raw = file_storage.read()
        try:
            content = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = raw.decode('latin-1')
        reader = csv.reader(io.StringIO(content))
        raw_headers = next(reader, [])
        headers = [_normalize_header(h) for h in raw_headers]
        for row in reader:
            # Asegurar que cada fila tenga la misma cantidad de columnas
            processed = [
                _sanitize_value(str(c).strip()) if c else ''
                for c in row
            ]
            # Si la fila está completamente vacía, saltarla
            if any(v for v in processed):
                rows.append(processed)

    elif ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_storage.read()), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, [])
        headers = [
            _normalize_header(str(h) if h is not None else '')
            for h in raw_headers
        ]
        for row in rows_iter:
            processed = [
                _sanitize_value(str(c).strip()) if c is not None else ''
                for c in row
            ]
            if any(v for v in processed):
                rows.append(processed)
        wb.close()

    # Normalizar: todas las filas deben tener len(headers) columnas
    ncols = len(headers)
    normalized_rows = []
    for row in rows:
        while len(row) < ncols:
            row.append('')
        normalized_rows.append(row[:ncols])
    rows = normalized_rows

    return {
        'headers': headers,
        'rows': rows,
        'total_rows': len(rows),
    }


def _build_field_map(headers):
    """
    Construye un mapa {campo_canónico: índice_en_headers}
    a partir de los headers normalizados.
    """
    field_map = {}
    for i, h in enumerate(headers):
        if h in _ALIAS_TO_FIELD:
            field_map[_ALIAS_TO_FIELD[h]] = i
    return field_map


def _get_value(row, field_map, field_name):
    """Extrae el valor de un campo desde una fila usando el field_map."""
    idx = field_map.get(field_name)
    if idx is None:
        return ''
    return row[idx] if idx < len(row) else ''


def _resolve_carrera(value):
    """Busca Carrera por código exacto o nombre (case-insensitive)."""
    if not value:
        return None
    # Por código exacto
    carrera = Carrera.query.filter_by(codigo=value).first()
    if carrera:
        return carrera
    # Por nombre (case-insensitive)
    carrera = Carrera.query.filter(Carrera.nombre.ilike(value)).first()
    return carrera


def _resolve_materia(value):
    """Busca Materia por código exacto o nombre (case-insensitive)."""
    if not value:
        return None
    materia = Materia.query.filter_by(codigo=value).first()
    if materia:
        return materia
    materia = Materia.query.filter(Materia.nombre.ilike(value)).first()
    return materia


def _resolve_alumno(numero_control):
    """Busca Alumno por numero_control."""
    if not numero_control:
        return None
    return Alumno.query.filter_by(numero_control=numero_control).first()


def _parse_bool(value):
    """Interpreta un valor como booleano."""
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ('true', '1', 'yes', 'sí', 'si', 'verdadero', 'pagada', 'pagado')


def _parse_float(value):
    """Convierte un valor a float, retorna 0.0 si falla."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _parse_int(value):
    """Convierte un valor a int, retorna 0 si falla."""
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0


# ============================================================
# PARSERS POR TIPO
# ============================================================

def _parse_alumnos(headers, rows):
    """
    Parsea filas de alumnos.
    Returns: {'rows': [...], 'errors': [...]}
    """
    field_map = _build_field_map(headers)
    result_rows = []
    errors = []

    # Verificar columnas requeridas
    required = ['numero_control', 'nombre', 'apellido_paterno', 'email', 'carrera']
    for req in required:
        if req not in field_map:
            errors.append({
                'row': 0,
                'field': req,
                'value': None,
                'message': f'Columna requerida "{req}" no encontrada en el archivo. '
                           f'Verifica que el archivo contenga una columna con ese nombre.'
            })
            return {'rows': [], 'errors': errors}

    for row_idx, row_values in enumerate(rows):
        row_num = row_idx + 1

        numero_control = _get_value(row_values, field_map, 'numero_control')
        nombre = _get_value(row_values, field_map, 'nombre')
        apellido_paterno = _get_value(row_values, field_map, 'apellido_paterno')
        apellido_materno = _get_value(row_values, field_map, 'apellido_materno')
        email = _get_value(row_values, field_map, 'email')
        carrera_val = _get_value(row_values, field_map, 'carrera')
        password = _get_value(row_values, field_map, 'password')

        # --- Validaciones ---
        row_errors = []

        if not numero_control:
            row_errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control, 'message': 'El número de control es requerido'
            })
        elif len(numero_control) > 20:
            row_errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control, 'message': 'El número de control no debe exceder 20 caracteres'
            })

        if not nombre:
            row_errors.append({
                'row': row_num, 'field': 'nombre',
                'value': nombre, 'message': 'El nombre es requerido'
            })

        if not apellido_paterno:
            row_errors.append({
                'row': row_num, 'field': 'apellido_paterno',
                'value': apellido_paterno, 'message': 'El apellido paterno es requerido'
            })

        if not email:
            row_errors.append({
                'row': row_num, 'field': 'email',
                'value': email, 'message': 'El email es requerido'
            })
        elif not EMAIL_REGEX.match(email):
            row_errors.append({
                'row': row_num, 'field': 'email',
                'value': email, 'message': 'Formato de email inválido'
            })
        elif len(email) > 120:
            row_errors.append({
                'row': row_num, 'field': 'email',
                'value': email, 'message': 'El email no debe exceder 120 caracteres'
            })

        if not carrera_val:
            row_errors.append({
                'row': row_num, 'field': 'carrera',
                'value': carrera_val, 'message': 'La carrera es requerida'
            })

        # Si ya hay errores en campos obligatorios, no seguir validando con DB
        if row_errors:
            errors.extend(row_errors)
            result_rows.append({
                'numero_control': numero_control,
                'nombre': nombre,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'email': email,
                'password': password or f"alumno{numero_control}",
                'carrera': carrera_val,
                'valid': False,
            })
            continue

        # Validar carrera
        carrera = _resolve_carrera(carrera_val)
        if not carrera:
            errors.append({
                'row': row_num, 'field': 'carrera',
                'value': carrera_val,
                'message': f'La carrera "{carrera_val}" no existe en el sistema'
            })
            result_rows.append({
                'numero_control': numero_control,
                'nombre': nombre,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'email': email,
                'password': password or f"alumno{numero_control}",
                'carrera': carrera_val,
                'valid': False,
            })
            continue

        # Validar unicidad de numero_control (contra DB)
        if Alumno.query.filter_by(numero_control=numero_control).first():
            errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control,
                'message': f'El número de control "{numero_control}" ya existe en el sistema'
            })
            result_rows.append({
                'numero_control': numero_control,
                'nombre': nombre,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'email': email,
                'password': password or f"alumno{numero_control}",
                'carrera': carrera_val,
                'carrera_id': carrera.id,
                'valid': False,
            })
            continue

        # Validar unicidad de email (contra DB)
        if Alumno.query.filter_by(email=email.lower()).first():
            errors.append({
                'row': row_num, 'field': 'email',
                'value': email,
                'message': f'El email "{email}" ya existe en el sistema'
            })
            result_rows.append({
                'numero_control': numero_control,
                'nombre': nombre,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'email': email,
                'password': password or f"alumno{numero_control}",
                'carrera': carrera_val,
                'carrera_id': carrera.id,
                'valid': False,
            })
            continue

        # Validar password (si se provee, mínimo 6 caracteres)
        final_password = password
        if password and len(password) < 6:
            errors.append({
                'row': row_num, 'field': 'password',
                'value': password,
                'message': 'La contraseña debe tener al menos 6 caracteres'
            })
            result_rows.append({
                'numero_control': numero_control,
                'nombre': nombre,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'email': email,
                'password': password,
                'carrera': carrera_val,
                'carrera_id': carrera.id,
                'valid': False,
            })
            continue

        if not password:
            final_password = f"alumno{numero_control}"

        # Fila válida
        row_data = {
            'numero_control': numero_control,
            'nombre': nombre,
            'apellido_paterno': apellido_paterno,
            'apellido_materno': apellido_materno if apellido_materno else None,
            'email': email.lower().strip(),
            'password': final_password,
            'carrera_id': carrera.id,
            'valid': True,
        }
        result_rows.append(row_data)

    return {'rows': result_rows, 'errors': errors}


def _parse_calificaciones(headers, rows):
    """
    Parsea filas de calificaciones.
    Returns: {'rows': [...], 'errors': [...]}
    """
    field_map = _build_field_map(headers)
    result_rows = []
    errors = []

    # Columnas requeridas
    required = ['numero_control', 'materia', 'calificacion_final', 'periodo', 'anio']
    for req in required:
        if req not in field_map:
            errors.append({
                'row': 0,
                'field': req,
                'value': None,
                'message': f'Columna requerida "{req}" no encontrada en el archivo'
            })
            return {'rows': [], 'errors': errors}

    for row_idx, row_values in enumerate(rows):
        row_num = row_idx + 1

        numero_control = _get_value(row_values, field_map, 'numero_control')
        materia_val = _get_value(row_values, field_map, 'materia')
        calificacion_final_str = _get_value(row_values, field_map, 'calificacion_final')
        periodo = _get_value(row_values, field_map, 'periodo')
        anio_str = _get_value(row_values, field_map, 'anio')

        # Campos opcionales
        practica_1_str = _get_value(row_values, field_map, 'practica_1')
        practica_2_str = _get_value(row_values, field_map, 'practica_2')
        extra_1_str = _get_value(row_values, field_map, 'extra_1')
        extra_2_str = _get_value(row_values, field_map, 'extra_2')
        asistencia_1_str = _get_value(row_values, field_map, 'asistencia_1')
        asistencia_2_str = _get_value(row_values, field_map, 'asistencia_2')
        asistencia_3_str = _get_value(row_values, field_map, 'asistencia_3')
        asistencia_4_str = _get_value(row_values, field_map, 'asistencia_4')
        asistencia_5_str = _get_value(row_values, field_map, 'asistencia_5')

        row_errors = []

        # Validar numero_control
        if not numero_control:
            row_errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control, 'message': 'El número de control es requerido'
            })

        # Validar materia
        if not materia_val:
            row_errors.append({
                'row': row_num, 'field': 'materia',
                'value': materia_val, 'message': 'La materia es requerida'
            })

        # Validar calificacion_final
        calificacion_final = _parse_float(calificacion_final_str)
        if not calificacion_final_str:
            row_errors.append({
                'row': row_num, 'field': 'calificacion_final',
                'value': calificacion_final_str, 'message': 'La calificación final es requerida'
            })
        elif calificacion_final < 0 or calificacion_final > 10:
            row_errors.append({
                'row': row_num, 'field': 'calificacion_final',
                'value': calificacion_final_str,
                'message': 'La calificación final debe estar entre 0 y 10'
            })

        # Validar periodo
        if not periodo:
            row_errors.append({
                'row': row_num, 'field': 'periodo',
                'value': periodo, 'message': 'El período es requerido'
            })

        # Validar anio
        anio = _parse_int(anio_str)
        if not anio_str:
            row_errors.append({
                'row': row_num, 'field': 'anio',
                'value': anio_str, 'message': 'El año es requerido'
            })
        elif anio < 1900 or anio > 2099:
            row_errors.append({
                'row': row_num, 'field': 'anio',
                'value': anio_str, 'message': 'El año debe estar entre 1900 y 2099'
            })

        # Si errores en campos clave, acumular y continuar
        if row_errors:
            errors.extend(row_errors)
            result_rows.append({
                'numero_control': numero_control,
                'materia': materia_val,
                'calificacion_final': calificacion_final_str,
                'periodo': periodo,
                'anio': anio_str,
                'valid': False,
            })
            continue

        # Resolver alumno
        alumno = _resolve_alumno(numero_control)
        if not alumno:
            errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control,
                'message': f'El alumno con número de control "{numero_control}" no existe'
            })
            result_rows.append({
                'numero_control': numero_control,
                'materia': materia_val,
                'calificacion_final': calificacion_final,
                'periodo': periodo,
                'anio': anio,
                'valid': False,
            })
            continue

        # Resolver materia
        materia = _resolve_materia(materia_val)
        if not materia:
            errors.append({
                'row': row_num, 'field': 'materia',
                'value': materia_val,
                'message': f'La materia "{materia_val}" no existe en el sistema'
            })
            result_rows.append({
                'numero_control': numero_control,
                'materia': materia_val,
                'calificacion_final': calificacion_final,
                'periodo': periodo,
                'anio': anio,
                'valid': False,
            })
            continue

        # Parsear campos opcionales
        def clamp_float(val_str, lo=0, hi=10):
            v = _parse_float(val_str)
            return max(lo, min(hi, v))

        def clamp_bool(val_str):
            return 1 if _parse_bool(val_str) else 0

        row_data = {
            'alumno_id': alumno.id,
            'materia_id': materia.id,
            'numero_control': numero_control,
            'calificacion_final': clamp_float(calificacion_final_str),
            'periodo': periodo,
            'anio': anio,
            'practica_1': clamp_float(practica_1_str),
            'practica_2': clamp_float(practica_2_str),
            'extra_1': clamp_float(extra_1_str),
            'extra_2': clamp_float(extra_2_str),
            'asistencia_1': clamp_bool(asistencia_1_str),
            'asistencia_2': clamp_bool(asistencia_2_str),
            'asistencia_3': clamp_bool(asistencia_3_str),
            'asistencia_4': clamp_bool(asistencia_4_str),
            'asistencia_5': clamp_bool(asistencia_5_str),
            'valid': True,
        }
        result_rows.append(row_data)

    return {'rows': result_rows, 'errors': errors}


def _parse_pagos(headers, rows):
    """
    Parsea filas de pagos (notas de remisión).
    Returns: {'rows': [...], 'errors': [...]}
    """
    field_map = _build_field_map(headers)
    result_rows = []
    errors = []

    # Columnas requeridas
    required = ['numero_control', 'concepto', 'monto']
    for req in required:
        if req not in field_map:
            errors.append({
                'row': 0,
                'field': req,
                'value': None,
                'message': f'Columna requerida "{req}" no encontrada en el archivo'
            })
            return {'rows': [], 'errors': errors}

    for row_idx, row_values in enumerate(rows):
        row_num = row_idx + 1

        numero_control = _get_value(row_values, field_map, 'numero_control')
        concepto = _get_value(row_values, field_map, 'concepto')
        monto_str = _get_value(row_values, field_map, 'monto')
        fecha_emision_str = _get_value(row_values, field_map, 'fecha_emision')
        pagada_str = _get_value(row_values, field_map, 'pagada')
        fecha_pago_str = _get_value(row_values, field_map, 'fecha_pago')

        row_errors = []

        # Validar numero_control
        if not numero_control:
            row_errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control, 'message': 'El número de control es requerido'
            })

        # Validar concepto
        if not concepto:
            row_errors.append({
                'row': row_num, 'field': 'concepto',
                'value': concepto, 'message': 'El concepto es requerido'
            })

        # Validar monto
        monto = _parse_float(monto_str)
        if not monto_str:
            row_errors.append({
                'row': row_num, 'field': 'monto',
                'value': monto_str, 'message': 'El monto es requerido'
            })
        elif monto <= 0:
            row_errors.append({
                'row': row_num, 'field': 'monto',
                'value': monto_str, 'message': 'El monto debe ser mayor a 0'
            })

        if row_errors:
            errors.extend(row_errors)
            result_rows.append({
                'numero_control': numero_control,
                'concepto': concepto,
                'monto': monto_str,
                'fecha_emision': fecha_emision_str,
                'pagada': pagada_str,
                'fecha_pago': fecha_pago_str,
                'valid': False,
            })
            continue

        # Resolver alumno
        alumno = _resolve_alumno(numero_control)
        if not alumno:
            errors.append({
                'row': row_num, 'field': 'numero_control',
                'value': numero_control,
                'message': f'El alumno con número de control "{numero_control}" no existe'
            })
            result_rows.append({
                'numero_control': numero_control,
                'concepto': concepto,
                'monto': monto,
                'valid': False,
            })
            continue

        # Parsear fecha_emision (default: today)
        fecha_emision = date.today()
        if fecha_emision_str:
            try:
                fecha_emision = datetime.strptime(fecha_emision_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                errors.append({
                    'row': row_num, 'field': 'fecha_emision',
                    'value': fecha_emision_str,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                })
                result_rows.append({
                    'numero_control': numero_control,
                    'concepto': concepto,
                    'monto': monto,
                    'valid': False,
                })
                continue

        # Parsear pagada (default: false)
        pagada = _parse_bool(pagada_str) if pagada_str else False

        # Parsear fecha_pago (solo si pagada=true)
        fecha_pago = None
        if pagada and fecha_pago_str:
            try:
                fecha_pago = datetime.strptime(fecha_pago_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                errors.append({
                    'row': row_num, 'field': 'fecha_pago',
                    'value': fecha_pago_str,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                })
                result_rows.append({
                    'numero_control': numero_control,
                    'concepto': concepto,
                    'monto': monto,
                    'valid': False,
                })
                continue
        elif pagada and not fecha_pago_str:
            fecha_pago = date.today()

        row_data = {
            'alumno_id': alumno.id,
            'numero_control': numero_control,
            'concepto': concepto.strip(),
            'monto': monto,
            'fecha_emision': fecha_emision,
            'pagada': pagada,
            'fecha_pago': fecha_pago,
            'valid': True,
        }
        result_rows.append(row_data)

    return {'rows': result_rows, 'errors': errors}


# ============================================================
# PARSER: CARRERAS
# ============================================================

def _parse_carreras(headers, rows):
    """
    Parsea filas de carreras.
    Columnas: codigo (req), nombre (req), descripcion (opc), activa (opc)
    Returns: {'rows': [...], 'errors': [...]}
    """
    field_map = _build_field_map(headers)
    result_rows = []
    errors = []

    required = ['codigo', 'nombre']
    for req in required:
        if req not in field_map:
            errors.append({
                'row': 0,
                'field': req,
                'value': None,
                'message': f'Columna requerida "{req}" no encontrada en el archivo'
            })
            return {'rows': [], 'errors': errors}

    for row_idx, row_values in enumerate(rows):
        row_num = row_idx + 1

        codigo = _get_value(row_values, field_map, 'codigo')
        nombre = _get_value(row_values, field_map, 'nombre')
        descripcion = _get_value(row_values, field_map, 'descripcion')
        activa_str = _get_value(row_values, field_map, 'activa')

        row_errors = []

        if not codigo:
            row_errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo, 'message': 'El código es requerido'
            })
        elif len(codigo) > 20:
            row_errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo, 'message': 'El código no debe exceder 20 caracteres'
            })

        if not nombre:
            row_errors.append({
                'row': row_num, 'field': 'nombre',
                'value': nombre, 'message': 'El nombre es requerido'
            })

        if row_errors:
            errors.extend(row_errors)
            result_rows.append({
                'codigo': codigo, 'nombre': nombre,
                'valid': False,
            })
            continue

        # Validar unicidad de codigo
        if Carrera.query.filter_by(codigo=codigo).first():
            errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo,
                'message': f'El código "{codigo}" ya existe en el sistema'
            })
            result_rows.append({
                'codigo': codigo, 'nombre': nombre,
                'valid': False,
            })
            continue

        activa = _parse_bool(activa_str) if activa_str else True

        row_data = {
            'codigo': codigo.upper().strip(),
            'nombre': nombre.strip(),
            'descripcion': descripcion.strip() if descripcion else None,
            'activa': activa,
            'valid': True,
        }
        result_rows.append(row_data)

    return {'rows': result_rows, 'errors': errors}


# ============================================================
# PARSER: MATERIAS
# ============================================================

def _parse_materias(headers, rows):
    """
    Parsea filas de materias.
    Columnas: codigo (req), nombre (req), carrera (req),
              creditos (opc, default 0)
    Returns: {'rows': [...], 'errors': [...]}
    """
    field_map = _build_field_map(headers)
    result_rows = []
    errors = []

    required = ['codigo', 'nombre', 'carrera']
    for req in required:
        if req not in field_map:
            errors.append({
                'row': 0,
                'field': req,
                'value': None,
                'message': f'Columna requerida "{req}" no encontrada en el archivo'
            })
            return {'rows': [], 'errors': errors}

    for row_idx, row_values in enumerate(rows):
        row_num = row_idx + 1

        codigo = _get_value(row_values, field_map, 'codigo')
        nombre = _get_value(row_values, field_map, 'nombre')
        carrera_val = _get_value(row_values, field_map, 'carrera')
        creditos_str = _get_value(row_values, field_map, 'creditos')

        row_errors = []

        if not codigo:
            row_errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo, 'message': 'El código es requerido'
            })
        elif len(codigo) > 20:
            row_errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo, 'message': 'El código no debe exceder 20 caracteres'
            })

        if not nombre:
            row_errors.append({
                'row': row_num, 'field': 'nombre',
                'value': nombre, 'message': 'El nombre es requerido'
            })

        if not carrera_val:
            row_errors.append({
                'row': row_num, 'field': 'carrera',
                'value': carrera_val, 'message': 'La carrera es requerida'
            })

        if row_errors:
            errors.extend(row_errors)
            result_rows.append({
                'codigo': codigo, 'nombre': nombre,
                'carrera': carrera_val,
                'valid': False,
            })
            continue

        # Validar carrera
        carrera = _resolve_carrera(carrera_val)
        if not carrera:
            errors.append({
                'row': row_num, 'field': 'carrera',
                'value': carrera_val,
                'message': f'La carrera "{carrera_val}" no existe en el sistema'
            })
            result_rows.append({
                'codigo': codigo, 'nombre': nombre,
                'carrera': carrera_val,
                'valid': False,
            })
            continue

        # Validar unicidad de codigo dentro de la misma carrera
        if Materia.query.filter_by(codigo=codigo, carrera_id=carrera.id).first():
            errors.append({
                'row': row_num, 'field': 'codigo',
                'value': codigo,
                'message': f'El código "{codigo}" ya existe para la carrera "{carrera.nombre}"'
            })
            result_rows.append({
                'codigo': codigo, 'nombre': nombre,
                'carrera': carrera_val,
                'carrera_id': carrera.id,
                'valid': False,
            })
            continue

        creditos = _parse_int(creditos_str) if creditos_str else 0

        row_data = {
            'codigo': codigo.upper().strip(),
            'nombre': nombre.strip(),
            'carrera_id': carrera.id,
            'creditos': creditos,
            'valid': True,
        }
        result_rows.append(row_data)

    return {'rows': result_rows, 'errors': errors}


# ============================================================
# ENDPOINT: PREVIEW
# ============================================================

@imports_bp.route('/preview', methods=['POST'])
@admin_required
def preview_import():
    """
    Previsualiza un archivo de importación.
    Lee headers + primeras filas, valida y retorna preview.
    """
    # Validar tipo
    tipo = request.form.get('tipo', '').strip().lower()
    if tipo not in ('alumnos', 'calificaciones', 'pagos', 'carreras', 'materias'):
        return jsonify({
            'error': 'Tipo de importación inválido. Use: alumnos, calificaciones, pagos, carreras o materias',
            'code': 'INVALID_TYPE'
        }), 400

    # Validar archivo
    if 'file' not in request.files:
        return jsonify({
            'error': 'Archivo requerido',
            'code': 'FILE_REQUIRED'
        }), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({
            'error': 'Archivo sin nombre',
            'code': 'EMPTY_FILE'
        }), 400

    # Validar extensión
    try:
        _detect_file_type(file.filename)
    except ValueError as e:
        return jsonify({
            'error': str(e),
            'code': 'INVALID_FILE_TYPE'
        }), 400

    # Validar tamaño
    if request.content_length and request.content_length > MAX_CONTENT_LENGTH:
        return jsonify({
            'error': 'El archivo excede el límite de 10MB',
            'code': 'FILE_TOO_LARGE'
        }), 413

    # Leer archivo
    try:
        data = _read_file(file)
    except Exception as e:
        return jsonify({
            'error': f'Error al leer el archivo: {str(e)}',
            'code': 'FILE_READ_ERROR'
        }), 400

    headers = data['headers']
    all_rows = data['rows']
    total_rows = len(all_rows)

    if total_rows == 0:
        return jsonify({
            'error': 'El archivo no contiene datos',
            'code': 'EMPTY_FILE'
        }), 400

    # Procesar solo las primeras 10 filas para preview
    preview_rows = all_rows[:10]

    # Definir columnas esperadas
    column_defs = {
        'alumnos': ['numero_control', 'nombre', 'apellido_paterno', 'apellido_materno', 'email', 'carrera'],
        'calificaciones': ['numero_control', 'materia', 'calificacion_final', 'periodo', 'anio'],
        'pagos': ['numero_control', 'concepto', 'monto', 'fecha_emision', 'pagada'],
        'carreras': ['codigo', 'nombre', 'descripcion', 'activa'],
        'materias': ['codigo', 'nombre', 'carrera', 'creditos'],
    }

    # Ejecutar parser correspondiente
    parsers = {
        'alumnos': _parse_alumnos,
        'calificaciones': _parse_calificaciones,
        'pagos': _parse_pagos,
        'carreras': _parse_carreras,
        'materias': _parse_materias,
    }

    parser_result = parsers[tipo](headers, preview_rows)

    # Separar errores estructurales (row=0) de errores por fila
    structural_errors = [e for e in parser_result['errors'] if e.get('row') == 0]
    importable = len(structural_errors) == 0

    # Construir rows_preview
    rows_preview = []
    row_errors_by_row = {}
    for err in parser_result['errors']:
        r = err.get('row', 0)
        if r > 0:
            if r not in row_errors_by_row:
                row_errors_by_row[r] = []
            row_errors_by_row[r].append(err)

    for i, row_data in enumerate(parser_result['rows']):
        row_num = i + 1
        row_errs = row_errors_by_row.get(row_num, [])
        rows_preview.append({
            'row': row_num,
            'data': row_data,
            'valid': row_data.get('valid', len(row_errs) == 0),
            'errors': row_errs,
        })

    # Construir warnings a partir de columnas no reconocidas
    field_map = _build_field_map(headers)
    recognized = set(field_map.keys())
    warnings = []
    for h in headers:
        if h and h not in _ALIAS_TO_FIELD:
            warnings.append(f"Columna '{h}' no será importada (no reconocida)")

    return jsonify({
        'columns': column_defs[tipo],
        'rows_preview': rows_preview,
        'total_rows': total_rows,
        'importable': importable,
        'warnings': warnings,
    }), 200


# ============================================================
# ENDPOINT: EXECUTE
# ============================================================

@imports_bp.route('/execute', methods=['POST'])
@admin_required
def execute_import():
    """
    Ejecuta la importación completa.
    Todas las filas se validan primero. Si hay errores, se retornan
    sin modificar la DB (rollback implícito). Si todo OK, transacción.
    """
    # Validar tipo
    tipo = request.form.get('tipo', '').strip().lower()
    if tipo not in ('alumnos', 'calificaciones', 'pagos', 'carreras', 'materias'):
        return jsonify({
            'error': 'Tipo de importación inválido. Use: alumnos, calificaciones, pagos, carreras o materias',
            'code': 'INVALID_TYPE'
        }), 400

    # Validar archivo
    if 'file' not in request.files:
        return jsonify({
            'error': 'Archivo requerido',
            'code': 'FILE_REQUIRED'
        }), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({
            'error': 'Archivo sin nombre',
            'code': 'EMPTY_FILE'
        }), 400

    try:
        _detect_file_type(file.filename)
    except ValueError as e:
        return jsonify({
            'error': str(e),
            'code': 'INVALID_FILE_TYPE'
        }), 400

    if request.content_length and request.content_length > MAX_CONTENT_LENGTH:
        return jsonify({
            'error': 'El archivo excede el límite de 10MB',
            'code': 'FILE_TOO_LARGE'
        }), 413

    # Leer archivo completo
    try:
        data = _read_file(file)
    except Exception as e:
        return jsonify({
            'error': f'Error al leer el archivo: {str(e)}',
            'code': 'FILE_READ_ERROR'
        }), 400

    headers = data['headers']
    all_rows = data['rows']

    if len(all_rows) == 0:
        return jsonify({
            'error': 'El archivo no contiene datos',
            'code': 'EMPTY_FILE'
        }), 400

    # Ejecutar parser correspondiente (todas las filas)
    parsers = {
        'alumnos': _parse_alumnos,
        'calificaciones': _parse_calificaciones,
        'pagos': _parse_pagos,
        'carreras': _parse_carreras,
        'materias': _parse_materias,
    }

    parser_result = parsers[tipo](headers, all_rows)

    # Si hay errores (estructurales o por fila), retornar sin escribir
    if parser_result['errors']:
        # Separar estructurales (row=0) para el conteo
        row_errors = [e for e in parser_result['errors'] if e.get('row', 0) > 0]
        return jsonify({
            'status': 'error',
            'imported': 0,
            'errors': parser_result['errors'],
            'total_rows': len(all_rows),
            'error_count': len(parser_result['errors']),
        }), 200

    # Obtener admin_id del JWT
    claims = get_jwt()
    admin_id = claims.get('id', 1)

    # Ejecutar transacción
    try:
        result = _execute_transaction(tipo, parser_result['rows'], admin_id)
        return jsonify(result), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'imported': 0,
            'errors': [{
                'row': 0,
                'field': None,
                'value': None,
                'message': f'Error interno durante la importación: {str(e)}'
            }],
            'total_rows': len(all_rows),
            'error_count': 1,
        }), 500


def _execute_transaction(tipo, rows, admin_id):
    """Ejecuta la transacción según el tipo, dentro de batches."""
    created = 0
    updated = 0
    generated_passwords = []

    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]

        for row_data in batch:
            if tipo == 'alumnos':
                alumno = Alumno(
                    numero_control=row_data['numero_control'],
                    nombre=row_data['nombre'],
                    apellido_paterno=row_data['apellido_paterno'],
                    apellido_materno=row_data.get('apellido_materno') or None,
                    email=row_data['email'],
                    carrera_id=row_data['carrera_id'],
                    activo=True,
                    fecha_registro=date.today(),
                )
                alumno.set_password(row_data['password'])
                db.session.add(alumno)
                created += 1

                # Registrar password generado (si se generó automáticamente)
                generated_passwords.append({
                    'numero_control': row_data['numero_control'],
                    'password': row_data['password'],
                })

            elif tipo == 'calificaciones':
                # Upsert: buscar por unique constraint
                existing = Calificacion.query.filter_by(
                    alumno_id=row_data['alumno_id'],
                    materia_id=row_data['materia_id'],
                    periodo=row_data['periodo'],
                    anio=row_data['anio'],
                ).first()

                if existing:
                    existing.calificacion_final = row_data['calificacion_final']
                    existing.practica_1 = row_data['practica_1']
                    existing.practica_2 = row_data['practica_2']
                    existing.extra_1 = row_data['extra_1']
                    existing.extra_2 = row_data['extra_2']
                    existing.asistencia_1 = row_data['asistencia_1']
                    existing.asistencia_2 = row_data['asistencia_2']
                    existing.asistencia_3 = row_data['asistencia_3']
                    existing.asistencia_4 = row_data['asistencia_4']
                    existing.asistencia_5 = row_data['asistencia_5']
                    updated += 1
                else:
                    calif = Calificacion(
                        alumno_id=row_data['alumno_id'],
                        materia_id=row_data['materia_id'],
                        calificacion_final=row_data['calificacion_final'],
                        practica_1=row_data['practica_1'],
                        practica_2=row_data['practica_2'],
                        extra_1=row_data['extra_1'],
                        extra_2=row_data['extra_2'],
                        asistencia_1=row_data['asistencia_1'],
                        asistencia_2=row_data['asistencia_2'],
                        asistencia_3=row_data['asistencia_3'],
                        asistencia_4=row_data['asistencia_4'],
                        asistencia_5=row_data['asistencia_5'],
                        periodo=row_data['periodo'],
                        anio=row_data['anio'],
                    )
                    db.session.add(calif)
                    created += 1

            elif tipo == 'pagos':
                nota = NotaRemision(
                    alumno_id=row_data['alumno_id'],
                    concepto=row_data['concepto'],
                    monto=row_data['monto'],
                    fecha_emision=row_data['fecha_emision'],
                    pagada=row_data['pagada'],
                    fecha_pago=row_data.get('fecha_pago'),
                    created_by_id=admin_id,
                )
                db.session.add(nota)
                created += 1

            elif tipo == 'carreras':
                carrera = Carrera(
                    codigo=row_data['codigo'],
                    nombre=row_data['nombre'],
                    descripcion=row_data.get('descripcion'),
                    activa=row_data['activa'],
                )
                db.session.add(carrera)
                created += 1

            elif tipo == 'materias':
                materia = Materia(
                    codigo=row_data['codigo'],
                    nombre=row_data['nombre'],
                    carrera_id=row_data['carrera_id'],
                    creditos=row_data['creditos'],
                )
                db.session.add(materia)
                created += 1

        # Commit batch
        db.session.commit()

    # Construir respuesta según tipo
    details = {}
    if tipo == 'alumnos':
        details = {'alumnos_importados': created}
    elif tipo == 'calificaciones':
        details = {
            'calificaciones_creadas': created,
            'calificaciones_actualizadas': updated,
        }
    elif tipo == 'pagos':
        details = {'pagos_creados': created}
    elif tipo == 'carreras':
        details = {'carreras_importadas': created}
    elif tipo == 'materias':
        details = {'materias_importadas': created}

    return {
        'status': 'success',
        'imported': created + updated,
        'created': created,
        'updated': updated,
        'errors': [],
        'generated_passwords': generated_passwords,
        'details': details,
    }
